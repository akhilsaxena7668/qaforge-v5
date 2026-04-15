"""QAForge Gemini — Report & Excel Generator"""

import json
from datetime import datetime
from pathlib import Path
from typing import Dict, Any

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ReportGenerator:
    def generate(self, result: Dict, suite: Dict, fmt: str, out_dir: Path) -> Path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        rid = result["run_id"][:8]
        if fmt == "json":
            return self._json(result, suite, out_dir, ts, rid)
        if fmt == "excel":
            return self._excel(result, suite, out_dir, ts, rid)
        return self._html(result, suite, out_dir, ts, rid)

    # ── JSON ─────────────────────────────────────────────────────────────────
    def _json(self, result, suite, out_dir, ts, rid):
        path = out_dir / f"report_{rid}_{ts}.json"
        path.write_text(json.dumps({"generated_at": datetime.now().isoformat(), "suite": suite, "result": result}, indent=2))
        return path

    # ── EXCEL ────────────────────────────────────────────────────────────────
    def _excel(self, result, suite, out_dir, ts, rid):
        if not EXCEL_AVAILABLE:
            raise RuntimeError("openpyxl not installed")
        path = out_dir / f"report_{rid}_{ts}.xlsx"
        wb = openpyxl.Workbook()

        # ── colours ────────────────────────────────────────
        C_DARK    = "0A1628"
        C_HEADER  = "0D2137"
        C_ACCENT  = "00B4D8"
        C_PASS    = "2DC653"
        C_FAIL    = "E63946"
        C_WARN    = "FF9F1C"
        C_CRIT    = "E63946"
        C_HIGH    = "FF6B35"
        C_MED     = "FF9F1C"
        C_LOW     = "6B7280"
        C_WHITE   = "FFFFFF"
        C_LIGHT   = "E8F4FD"

        def hdr_font(sz=11, bold=True, color=C_WHITE):
            return Font(name="Calibri", size=sz, bold=bold, color=color)
        def fill(hex_):
            return PatternFill("solid", fgColor=hex_)
        def thin_border():
            s = Side(style="thin", color="2A3F55")
            return Border(left=s, right=s, top=s, bottom=s)
        def center():
            return Alignment(horizontal="center", vertical="center", wrap_text=True)
        def left():
            return Alignment(horizontal="left", vertical="center", wrap_text=True)

        summary = result.get("summary", {})
        tests   = result.get("tests", [])
        total   = summary.get("total", len(tests))
        passed  = summary.get("passed", 0)
        failed  = summary.get("failed", 0)
        rate    = summary.get("pass_rate", 0)
        suite_name = suite.get("name", "Test Suite")

        # ═══ SHEET 1: SUMMARY ════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.sheet_view.showGridLines = False
        ws1.column_dimensions["A"].width = 3
        ws1.column_dimensions["B"].width = 28
        ws1.column_dimensions["C"].width = 22
        ws1.column_dimensions["D"].width = 22
        ws1.column_dimensions["E"].width = 22
        ws1.column_dimensions["F"].width = 22
        ws1.row_dimensions[1].height = 8

        # Title band
        for r in range(2, 7):
            ws1.row_dimensions[r].height = 16
        ws1.merge_cells("B2:F6")
        tc = ws1["B2"]
        tc.value = f"QAForge — Test Report"
        tc.font  = Font(name="Calibri", size=22, bold=True, color=C_ACCENT)
        tc.fill  = fill(C_DARK)
        tc.alignment = center()

        ws1.merge_cells("B7:F8")
        sn = ws1["B7"]
        sn.value = suite_name
        sn.font  = Font(name="Calibri", size=13, bold=True, color=C_WHITE)
        sn.fill  = fill(C_HEADER)
        sn.alignment = center()
        ws1.row_dimensions[7].height = 20
        ws1.row_dimensions[8].height = 14

        # Meta row
        meta = [
            ("Run ID",      result.get("run_id","")[:16]),
            ("App Type",    suite.get("app_type","").upper()),
            ("Environment", result.get("environment","staging").upper()),
            ("AI Model",    suite.get("model_used","Gemini")),
            ("Generated",   datetime.now().strftime("%Y-%m-%d %H:%M")),
        ]
        ws1.row_dimensions[9].height = 14
        ws1.row_dimensions[10].height = 28
        for col_i, (lbl, val) in enumerate(meta):
            col = col_i + 2  # B..F
            cell_lbl = ws1.cell(row=9,  column=col, value=lbl)
            cell_val = ws1.cell(row=10, column=col, value=val)
            cell_lbl.font = Font(name="Calibri", size=8, bold=True, color="8BAFC7")
            cell_lbl.fill = fill(C_HEADER)
            cell_lbl.alignment = center()
            cell_val.font = Font(name="Calibri", size=10, bold=True, color=C_WHITE)
            cell_val.fill = fill(C_DARK)
            cell_val.alignment = center()
            cell_val.border = thin_border()

        ws1.row_dimensions[11].height = 10

        # KPI cards
        kpis = [
            ("TOTAL",   total,          C_ACCENT, "B"),
            ("PASSED",  passed,         C_PASS,   "C"),
            ("FAILED",  failed,         C_FAIL,   "D"),
            ("PASS RATE", f"{rate}%",   C_ACCENT, "E"),
            ("DURATION", f"{sum(t.get('duration_ms',0) for t in tests)//1000}s", C_WARN, "F"),
        ]
        for lbl, val, color, col in kpis:
            for row_off, (r_val, r_h) in enumerate([(lbl,16),(str(val),36),(None,8)]):
                row = 12 + row_off
                ws1.row_dimensions[row].height = r_h
                c = ws1[f"{col}{row}"]
                c.fill = fill(C_DARK if row_off == 1 else C_HEADER)
                c.alignment = center()
                if r_val:
                    c.value = r_val
                    if row_off == 0:
                        c.font = Font(name="Calibri", size=9, bold=True, color=color)
                    else:
                        c.font = Font(name="Calibri", size=26, bold=True, color=color)
                c.border = thin_border()

        # Category breakdown table
        cat_counts: Dict[str, Dict] = {}
        for t in tests:
            cat = t.get("category","other")
            if cat not in cat_counts:
                cat_counts[cat] = {"total":0,"pass":0,"fail":0}
            cat_counts[cat]["total"] += 1
            if t["status"] == "pass":
                cat_counts[cat]["pass"] += 1
            else:
                cat_counts[cat]["fail"] += 1

        start_row = 17
        ws1.merge_cells(f"B{start_row}:F{start_row}")
        th = ws1[f"B{start_row}"]
        th.value = "RESULTS BY CATEGORY"
        th.font  = Font(name="Calibri", size=10, bold=True, color=C_ACCENT)
        th.fill  = fill(C_HEADER)
        th.alignment = left()
        th.border = thin_border()
        ws1.row_dimensions[start_row].height = 22

        hdrs = ["Category","Total","Passed","Failed","Pass Rate"]
        for ci, h in enumerate(hdrs):
            c = ws1.cell(row=start_row+1, column=ci+2, value=h)
            c.font  = hdr_font(9)
            c.fill  = fill("152232")
            c.alignment = center()
            c.border = thin_border()
        ws1.row_dimensions[start_row+1].height = 20

        for ri, (cat, vals) in enumerate(cat_counts.items()):
            row = start_row + 2 + ri
            ws1.row_dimensions[row].height = 18
            pr = round(vals["pass"]/vals["total"]*100,1) if vals["total"] else 0
            row_data = [cat.upper(), vals["total"], vals["pass"], vals["fail"], f"{pr}%"]
            alt = fill("0E1C2E") if ri%2==0 else fill(C_DARK)
            for ci, val in enumerate(row_data):
                c = ws1.cell(row=row, column=ci+2, value=val)
                c.fill = alt
                c.font = Font(name="Calibri", size=9, color=C_WHITE)
                c.alignment = center()
                c.border = thin_border()
                if ci == 3:  # Failed
                    c.font = Font(name="Calibri", size=9, color=C_FAIL if val>0 else C_WHITE, bold=val>0)

        # ═══ SHEET 2: ALL TESTS ══════════════════════════════════════════════
        ws2 = wb.create_sheet("Test Results")
        ws2.sheet_view.showGridLines = False
        col_widths = [3, 10, 38, 16, 12, 12, 10, 48, 18]
        col_letters = ["A","B","C","D","E","F","G","H","I"]
        for l, w in zip(col_letters, col_widths):
            ws2.column_dimensions[l].width = w

        ws2.merge_cells("B1:I1")
        h = ws2["B1"]
        h.value = f"DETAILED TEST RESULTS — {suite_name}"
        h.font  = Font(name="Calibri", size=14, bold=True, color=C_ACCENT)
        h.fill  = fill(C_DARK)
        h.alignment = center()
        ws2.row_dimensions[1].height = 30

        headers = ["Test ID","Test Name","Category","Priority","Status","Duration","Notes","Executed At"]
        for ci, hdr in enumerate(headers):
            c = ws2.cell(row=2, column=ci+2, value=hdr)
            c.font  = hdr_font(9)
            c.fill  = fill(C_HEADER)
            c.alignment = center()
            c.border = thin_border()
        ws2.row_dimensions[2].height = 22

        for ri, t in enumerate(tests):
            row = ri + 3
            ws2.row_dimensions[row].height = 20
            is_fail = t["status"] == "fail"
            alt = fill("0E1C2E") if ri%2==0 else fill(C_DARK)
            row_vals = [
                t.get("test_id",""),
                t.get("test_name",""),
                t.get("category","").upper(),
                t.get("priority","").upper(),
                ("✓ PASS" if not is_fail else "✗ FAIL"),
                f"{t.get('duration_ms',0)} ms",
                t.get("note",""),
                t.get("executed_at",""),
            ]
            for ci, val in enumerate(row_vals):
                c = ws2.cell(row=row, column=ci+2, value=val)
                c.fill = alt
                c.border = thin_border()
                c.alignment = center() if ci not in [1,6] else left()
                if ci == 4:
                    c.font = Font(name="Calibri", size=9, bold=True,
                                  color=C_PASS if not is_fail else C_FAIL)
                elif ci == 3:
                    pmap = {"CRITICAL":C_CRIT,"HIGH":C_HIGH,"MEDIUM":C_MED,"LOW":C_LOW}
                    c.font = Font(name="Calibri", size=9, bold=True,
                                  color=pmap.get(str(val), C_WHITE))
                else:
                    c.font = Font(name="Calibri", size=9, color=C_WHITE)

        # ═══ SHEET 3: CHARTS ════════════════════════════════════════════════
        ws3 = wb.create_sheet("Charts")
        ws3.sheet_view.showGridLines = False
        ws3["B2"].value = "Category"
        ws3["C2"].value = "Passed"
        ws3["D2"].value = "Failed"
        ws3["B2"].font = hdr_font(9); ws3["C2"].font = hdr_font(9); ws3["D2"].font = hdr_font(9)
        for c in ["B2","C2","D2"]:
            ws3[c].fill = fill(C_HEADER); ws3[c].alignment = center()

        for ri, (cat, vals) in enumerate(cat_counts.items()):
            ws3.cell(row=3+ri, column=2, value=cat.upper())
            ws3.cell(row=3+ri, column=3, value=vals["pass"])
            ws3.cell(row=3+ri, column=4, value=vals["fail"])

        # Bar chart
        bar = BarChart()
        bar.type = "col"
        bar.title = "Test Results by Category"
        bar.y_axis.title = "Count"
        bar.x_axis.title = "Category"
        bar.style = 10
        bar.width = 20; bar.height = 12
        n = len(cat_counts)
        data_ref   = Reference(ws3, min_col=3, max_col=4, min_row=2, max_row=2+n)
        cats_ref   = Reference(ws3, min_col=2, min_row=3, max_row=2+n)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        ws3.add_chart(bar, "F2")

        # Pie chart
        pie = PieChart()
        pie.title = "Overall Pass vs Fail"
        pie.style = 10
        pie.width = 14; pie.height = 12
        ws3["B20"].value = "Status"; ws3["C20"].value = "Count"
        ws3["B21"].value = "Passed"; ws3["C21"].value = passed
        ws3["B22"].value = "Failed"; ws3["C22"].value = failed
        pie_data = Reference(ws3, min_col=3, min_row=20, max_row=22)
        pie_cats = Reference(ws3, min_col=2, min_row=21, max_row=22)
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(pie_cats)
        slices = [DataPoint(idx=0), DataPoint(idx=1)]
        ws3.add_chart(pie, "F18")

        # Priority breakdown
        ws3["B30"].value = "Priority"; ws3["C30"].value = "Count"
        pri_map: Dict[str,int] = {}
        for t in tests:
            p = t.get("priority","medium")
            pri_map[p] = pri_map.get(p,0)+1
        for ri, (p,v) in enumerate(pri_map.items()):
            ws3.cell(row=31+ri, column=2, value=p.upper())
            ws3.cell(row=31+ri, column=3, value=v)
        pri_pie = PieChart()
        pri_pie.title = "Tests by Priority"
        pri_pie.style = 10
        pri_pie.width = 14; pri_pie.height = 12
        p_data = Reference(ws3, min_col=3, min_row=30, max_row=30+len(pri_map))
        p_cats = Reference(ws3, min_col=2, min_row=31, max_row=30+len(pri_map))
        pri_pie.add_data(p_data, titles_from_data=True)
        pri_pie.set_categories(p_cats)
        ws3.add_chart(pri_pie, "R2")

        wb.save(path)
        return path

    # ── HTML ─────────────────────────────────────────────────────────────────
    def _html(self, result, suite, out_dir, ts, rid):
        summary = result.get("summary", {})
        tests   = result.get("tests", [])
        total   = summary.get("total", len(tests))
        passed  = summary.get("passed", 0)
        failed  = summary.get("failed", 0)
        rate    = summary.get("pass_rate", 0)
        suite_name = suite.get("name", "Test Suite")
        rows = ""
        for t in tests:
            icon = "✓" if t["status"] == "pass" else "✗"
            cls  = "pass" if t["status"] == "pass" else "fail"
            pri  = t.get("priority","medium")
            rows += f"""<tr class="{cls}">
              <td><span class="badge b-{pri}">{pri.upper()}</span></td>
              <td class="mono">{t.get('test_id','')}</td>
              <td>{t.get('test_name','')}</td>
              <td class="mono">{t.get('category','').upper()}</td>
              <td class="st"><span class="sico">{icon}</span>{t['status'].upper()}</td>
              <td class="mono">{t.get('duration_ms',0)}ms</td>
              <td class="mono note">{t.get('note','')}</td>
            </tr>"""
        html = f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>QAForge Report — {suite_name}</title>
<link href="https://fonts.googleapis.com/css2?family=Share+Tech+Mono&family=Barlow:wght@400;700;900&display=swap" rel="stylesheet"/>
<style>
:root{{--bg:#060a0e;--bg2:#0d1520;--s:#111c28;--b:#1a2d40;--a:#00e5ff;--p:#2dc653;--f:#e63946;--t:#ddeaf8;--t2:#6f8ea8;--t3:#344f66;}}
*{{box-sizing:border-box;margin:0;padding:0}} body{{background:var(--bg);color:var(--t);font-family:'Barlow',sans-serif;padding:2rem}}
header{{border-bottom:2px solid var(--a);padding-bottom:1.5rem;margin-bottom:2rem}}
.logo{{font-size:1.1rem;font-weight:900;color:var(--a);letter-spacing:.08em;font-family:'Share Tech Mono',monospace}}
.sn{{font-size:2rem;font-weight:900;margin:.4rem 0}}
.meta{{font-family:'Share Tech Mono',monospace;font-size:.72rem;color:var(--t2);margin-top:.3rem}}
.kpis{{display:grid;grid-template-columns:repeat(5,1fr);gap:1rem;margin:1.5rem 0}}
.kpi{{background:var(--bg2);border:1px solid var(--b);padding:1rem;text-align:center}}
.kn{{font-size:2rem;font-weight:900;line-height:1}} .kl{{font-family:'Share Tech Mono',monospace;font-size:.6rem;color:var(--t3);margin-top:3px;text-transform:uppercase;letter-spacing:.1em}}
.ca{{color:var(--a)}} .cp{{color:var(--p)}} .cf{{color:var(--f)}} .cw{{color:#ff9f1c}}
table{{width:100%;border-collapse:collapse;font-size:.8rem;margin-top:1.5rem}}
thead{{background:var(--s)}}
th{{font-family:'Share Tech Mono',monospace;font-size:.62rem;letter-spacing:.1em;text-transform:uppercase;color:var(--t2);padding:.7rem .9rem;text-align:left;border-bottom:1px solid var(--b)}}
td{{padding:.6rem .9rem;border-bottom:1px solid var(--b);vertical-align:top}}
tr.pass td{{border-left:3px solid var(--p)}} tr.fail td{{border-left:3px solid var(--f)}}
.st{{font-weight:700}} tr.pass .sico{{color:var(--p)}} tr.fail .sico{{color:var(--f)}} .sico{{margin-right:4px}}
.mono{{font-family:'Share Tech Mono',monospace;font-size:.7rem}} .note{{color:var(--t2)}}
.badge{{font-family:'Share Tech Mono',monospace;font-size:.58rem;padding:2px 5px;font-weight:700}}
.b-critical{{background:#e63946;color:#000}} .b-high{{background:#ff6b35;color:#000}}
.b-medium{{background:#ff9f1c;color:#000}} .b-low{{background:#344f66;color:var(--t)}}
footer{{margin-top:2rem;font-family:'Share Tech Mono',monospace;font-size:.68rem;color:var(--t3);text-align:center;border-top:1px solid var(--b);padding-top:1rem}}
</style></head><body>
<header>
  <div class="logo">◈ QAFORGE — POWERED BY GEMINI AI</div>
  <div class="sn">{suite_name}</div>
  <div class="meta">Run: {result['run_id']} &nbsp;|&nbsp; {result.get('started_at','')} &nbsp;|&nbsp; {suite.get('model_used','Gemini')} &nbsp;|&nbsp; Env: {result.get('environment','staging').upper()}</div>
</header>
<div class="kpis">
  <div class="kpi"><div class="kn ca">{total}</div><div class="kl">Total</div></div>
  <div class="kpi"><div class="kn cp">{passed}</div><div class="kl">Passed</div></div>
  <div class="kpi"><div class="kn cf">{failed}</div><div class="kl">Failed</div></div>
  <div class="kpi"><div class="kn ca">{rate}%</div><div class="kl">Pass Rate</div></div>
  <div class="kpi"><div class="kn cw">{sum(t.get('duration_ms',0) for t in tests)//1000}s</div><div class="kl">Total Time</div></div>
</div>
<table><thead><tr><th>Priority</th><th>ID</th><th>Name</th><th>Category</th><th>Status</th><th>Duration</th><th>Notes</th></tr></thead>
<tbody>{rows}</tbody></table>
<footer>QAForge Gemini AI — Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</footer>
</body></html>"""
        path = out_dir / f"report_{rid}_{ts}.html"
        path.write_text(html)
        return path
